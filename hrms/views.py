from django.shortcuts import   render,redirect,get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login,get_user_model,logout
from django.http import JsonResponse,Http404,HttpResponse
from manager.models import Employee,Department,Designation,Grade,Position,Role, Permission, ShiftType,RolePermission,Holiday,Leave,Shift,Schedule,Attendance,LeaveApplication,AdvanceSalaryApplication,Salary,MonthlySalary
from django.db.models import Count
from django.views.decorators.csrf import csrf_exempt
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd
from io import BytesIO

from django.utils.dateparse import parse_duration

from django.http import HttpResponseForbidden
from functools import wraps


def has_permission(permission_name):
    print(f"Registering decorator for: {permission_name}") 
    def decorator(view_func):
        @wraps(view_func)
      
        def _wrapped_view(request, *args, **kwargs):
 
            if not request.user.is_authenticated:
                return redirect('login') 


            if request.user.is_superuser:
                return view_func(request, *args, **kwargs)

            try:
                employee = request.user.employee
                
            except Employee.DoesNotExist:
                return HttpResponseForbidden("No employee record linked to your account.")
            
            if not employee.role:
                return HttpResponseForbidden("No role assigned.")

            if RolePermission.objects.filter(role=employee.role, permission__name=permission_name).exists():
                return view_func(request, *args, **kwargs)

            return HttpResponseForbidden("You don't have permission to access this page.")
        return _wrapped_view
    return decorator



from django.db.models import Sum, Count



@login_required
@has_permission("view_dashboard")
def index(request):
    if not request.user.is_superuser:
        return redirect("/emp")

    today = date.today()
    current_month = today.month
    current_year = today.year

    # --- Employees ---
    total_employees = Employee.objects.count()

    monthly_employees = Employee.objects.filter(employee_type='monthly').count()
    daily_employees = Employee.objects.filter(employee_type='daily').count()

    # --- Attendance ---
    monthly_attendance_count = Attendance.objects.filter(
        date__month=current_month,
        date__year=current_year
    ).count()

    today_attendance = Attendance.objects.filter(date=today)

    today_attendance_overview = {
        "full_day": today_attendance.filter(type="Full Day").count(),
        "half_day": today_attendance.filter(type="Half Day").count(),
        "late": today_attendance.filter(type="Late").count(),
        "absent": today_attendance.filter(type="Absent").count(),
    }

    # --- Salary ---
    total_salary_paid = MonthlySalary.objects.filter(
        month=current_month,
        year=current_year
    ).aggregate(total=Sum('net_salary'))['total'] or 0

    # --- Advance Salary ---
    total_advance_salary = AdvanceSalaryApplication.objects.filter(
        superadmin_approval=True
    ).aggregate(total=Sum('amount_requested'))['total'] or 0

    # --- Masters ---
    total_departments = Department.objects.count()
    total_designations = Designation.objects.count()
    total_holidays = Holiday.objects.count()
    recent_employees = Employee.objects.all().select_related('department')[:5]


    context = {
        "dashboard": "active",

        "total_employees": total_employees,
        "monthly_employees": monthly_employees,
        "daily_employees": daily_employees,

        "monthly_attendance_count": monthly_attendance_count,
        "today_attendance": today_attendance_overview,

        "total_salary_paid": round(total_salary_paid),
        "total_advance_salary": round(total_advance_salary),

        "total_departments": total_departments,
        "total_designations": total_designations,
        "total_holidays": total_holidays,
        "recent_employees":recent_employees
    }

    dept_employee_data = (
    Employee.objects
    .values('department__name')
    .annotate(total=Count('id'))
    .order_by('department__name')
)

    context["dept_employee_data"] = list(dept_employee_data)

    return render(request, "index.html", context)

@login_required
def emp(request):
    today = date.today()
    current_month = today.month
    current_year = today.year

    # Employees under current user
    employees = Employee.objects.filter(under=request.user)

    total_employees = employees.count()

    monthly_employees = employees.filter(employee_type='monthly').count()
    daily_employees = employees.filter(employee_type='daily').count()

    # Attendance (current month)
    monthly_attendance_count = Attendance.objects.filter(
        employee__in=employees,
        date__month=current_month,
        date__year=current_year
    ).count()

    # Today's attendance
    today_attendance = Attendance.objects.filter(
        employee__in=employees,
        date=today
    )

    today_attendance_overview = {
        "full_day": today_attendance.filter(type="Full Day").count(),
        "half_day": today_attendance.filter(type="Half Day").count(),
        "late": today_attendance.filter(type="Late").count(),
        "absent": today_attendance.filter(type="Absent").count(),
    }

    # Salary paid (current month)
    total_salary_paid = MonthlySalary.objects.filter(
        employee__in=employees,
        month=current_month,
        year=current_year
    ).aggregate(total=Sum('net_salary'))['total'] or 0

    # Advance Salary (approved only)
    total_advance_salary = AdvanceSalaryApplication.objects.filter(
        employee__in=employees,
        superadmin_approval=True
    ).aggregate(total=Sum('amount_requested'))['total'] or 0
    total_departments = Department.objects.count()
    total_designations = Designation.objects.count()
    total_holidays = Holiday.objects.count()
    recent_employees = employees.select_related('department')[:5]


    context = {
        "dashboard": "active",

        "total_employees": total_employees,
        "monthly_employees": monthly_employees,
        "daily_employees": daily_employees,
        "recent_employees": recent_employees,


        "monthly_attendance_count": monthly_attendance_count,
        "today_attendance": today_attendance_overview,

        "total_salary_paid": total_salary_paid,
        "total_advance_salary": total_advance_salary,
           "total_departments": total_departments,
        "total_designations": total_designations,
        "total_holidays": total_holidays,
    }
    dept_employee_data = (
    Employee.objects
    .filter(under=request.user)
    .values('department__name')
    .annotate(total=Count('id'))
    .order_by('department__name')
    
)

    context["dept_employee_data"] = list(dept_employee_data)
 


    return render(request, "emp.html", context)

# @login_required
# def emp(request):

#     d = {'dashboard': "active"}
#     return render(request, "emp.html", d)





def login_(request):
    if request.user.is_superuser:
        return redirect("/")
    if request.user.is_authenticated:
        return redirect("/emp")
    return render(request,"login.html")






@login_required()
@has_permission("Employee Add")
def employeelist(request):
    from django.utils.timezone import now
    from datetime import timedelta

    today = now().date()
    last_30_days = today - timedelta(days=30)

    total_employees = Employee.objects.count()
    active_employees = Employee.objects.filter(status="Active").count()
    inactive_employees = Employee.objects.filter(status="Inactive").count()
    new_joiners = Employee.objects.filter(joining_date__gte=last_30_days).count()
  
    
 
    # employees = Employee.objects.select_related(
    #     "designation", "role", "department", "grade", "position"
    # ).all().order_by("-id")  # latest first
    if request.user.is_superuser:
        employees = Employee.objects.select_related(
            "designation", "role", "department", "grade", "position"
        ).all().order_by("-id")
    else:
        employees = Employee.objects.select_related(
            "designation", "role", "department", "grade", "position"
        ).filter(under=request.user).exclude(user=request.user).order_by("-id")

    d = {
        'hrmslist': "active",
        'roles': Role.objects.order_by('name'),
        'departments': Department.objects.filter(status="Active").order_by('name'),
        'grades': Grade.objects.filter(status="Active").order_by('name'),
        'positions': Position.objects.filter(status="Active").order_by('name'),
        'designations': Designation.objects.filter(status="Active").select_related("department").order_by('name'),
        "allusers" : Employee.objects.filter(
    role__role_permissions__isnull=False
).distinct(),
        # employee counts
        'total_employees': total_employees,
        'active_employees': active_employees,
        'inactive_employees': inactive_employees,
        'new_joiners': new_joiners,
          'employees': employees,
    }

    return render(request, "emoloyeelist.html", d)

@login_required()
def employeegird(request):
    d={'hrmsgrid':"active"}
    return render(request,"emoloyeegrid.html",d)

@login_required()
def edetails(request,data):
    emp=Employee.objects.get(id=data)
    grade=Grade.objects.all()
    position=Position.objects.all()
    department=Department.objects.all()
    designation=Designation.objects.all()
    conveyance=(emp.monthly_salary * emp.conveyance) // 100
    hra=(emp.monthly_salary * emp.hra) // 100
    special=(emp.monthly_salary * emp.special) // 100
    ta=(emp.monthly_salary * emp.ta) // 100
    da=(emp.monthly_salary * emp.da) // 100
    epfo_f=(emp.monthly_salary * emp.epfo / 100)
    esic_f=((emp.monthly_salary +hra+special+ta+da+conveyance) * emp.esic / 100)
   

    totalsalary=emp.monthly_salary+conveyance+hra+special+ta+da -epfo_f -esic_f

    d={'edetails':"active","details":emp,"conveyance":conveyance,"hra":hra,"special":special,"ta":ta,"da":da,"epfo_f":epfo_f,"esic_f":esic_f,"totalsalary":totalsalary,"grade":grade,"position":position,"department":department,"designation":designation}
    return render(request,"edetails.html",d)

@login_required
def department(request):
    departments = Department.objects.annotate(employee_count=Count('employee'))
    context = {
        "department": "active",
        "departments": departments,
    }
    return render(request, "department.html", context)


@login_required()
def ticket(request):
    d={'ticket':"active"}
    return render(request,"ticket.html",d)

@login_required()
def ticketdetail(request):
    d={'ticketdetail':"active"}
    return render(request,"ticketdetails.html",d)

@login_required()
def holiday(request):
    allholiday=Holiday.objects.all()
    d={'holiday':"active",'allholiday':allholiday}
    return render(request,"holiday.html",d)



@login_required()
def leave(request):
    alla = Leave.objects.filter(status="Active")
  


    from django.db.models import Q

    if request.user.is_superuser:
        ewew = LeaveApplication.objects.all()
    else:
        ewew = LeaveApplication.objects.filter(
            Q(employee__user=request.user) | Q(employee__under=request.user)
        )

    leave_data = []
    for leave in ewew:
        if leave.duration_type == 'full' and leave.from_date and leave.to_date:
            duration = (leave.to_date - leave.from_date).days + 1
        else:
            duration = 0.5  # For half day

        leave_data.append({
            "leave": leave,
            "duration": duration,
        })

    context = {
        "leave_status": "active",
        "alla": alla,
        "leave_applications": leave_data,
    }

    return render(request, "leave.html", context)




@login_required
def request_advance_salary(request):
    if request.method == "POST":
        try:
            employee = get_object_or_404(Employee, user=request.user)

            amount = request.POST.get("amount_requested")
            salary_month_raw = request.POST.get("salary_month")  # format: mm/yyyy
            repayment_method = request.POST.get("repayment_method")
            emi_months = request.POST.get("emi_months")
            reason = request.POST.get("reason")
            attachment = request.FILES.get("attachment")

            # Validation
            if not amount or not salary_month_raw or not repayment_method or not reason:
                return JsonResponse({"msg": "Required fields missing"}, status=400)

            # Convert salary month "mm/yyyy" → DateField (first day of month)
            try:
                salary_month = datetime.strptime(salary_month_raw, "%m/%Y").date().replace(day=1)
            except:
                return JsonResponse({"msg": "Invalid salary month format"}, status=400)

            # If single installment → emi null
            if repayment_method == "single":
                emi_months = None

            application = AdvanceSalaryApplication.objects.create(
                employee=employee,
                amount_requested=amount,
                salary_month=salary_month,
                repayment_method=repayment_method,
                emi_months=emi_months,
                reason=reason,
                attachment=attachment,
            )

            return JsonResponse({
                "msg": "Advance Salary Request Submitted Successfully",
                "id": application.id
            })

        except Exception as e:
            return JsonResponse({"msg": str(e)}, status=500)

    return JsonResponse({"msg": "Invalid Request Method"}, status=405)


from django.db.models import Q


@login_required()
def advancesalary(request):
    # Superadmin gets all  
    if request.user.is_superuser:
        applications = AdvanceSalaryApplication.objects.all()
     
    else:
        applications = AdvanceSalaryApplication.objects.filter(
            Q(employee__user=request.user) | Q(employee__under=request.user)
        )
  
    # Prepare Data List
    salary_data = []
    for app in applications:
        salary_data.append({
            "application": app,
            "id": app.id,
            "amount": app.amount_requested,
            "salary_month": app.salary_month,
            "repayment_method": app.repayment_method,
            "emi_months": app.emi_months,
            "admin_approval": app.admin_approval,
            "superadmin_approval": app.superadmin_approval,
        })

    context = {
        "advancesalary": "active",
        "applications": salary_data,
    }

    

    return render(request, "advancesalary.html", context)

@login_required
def delete_advance_salary(request, id):
 
    obj = get_object_or_404(AdvanceSalaryApplication, id=id)
    obj.delete()
   
    return redirect("/advancesalary")  # change to your actual page



@login_required()
def leavesetting(request):
    alll=Leave.objects.all()
    d={'leavesetting':"active",'alll':alll}
    return render(request,"leavesetting.html",d)

# @login_required()
# def attendence(request,data):
#     currentdata=Employee.objects.get(id=data)
#     d={'attendence':"active","currentdata":currentdata}
#     return render(request,"attendence.html",d)
from django.utils.timezone import localdate, now
import calendar

@login_required()

def attendence(request, data=None):
    if not data:
        ddsd=User.objects.get(username=request.user).email
        currentdata = Employee.objects.get(email=ddsd)
    else:
        currentdata = Employee.objects.get(id=data)

    
    

    # Get selected month and year
    selected_month = request.GET.get("month")
    selected_year = request.GET.get("year")

    try:
        month = int(selected_month) if selected_month else localdate().month
        year = int(selected_year) if selected_year else localdate().year
    except ValueError:
        month = localdate().month
        year = localdate().year

    # Set 'today' based on selected year/month
    today = localdate().replace(year=year, month=month, day=1)

    # Calculate month start and end
    month_start = today.replace(day=1)
    last_day = calendar.monthrange(today.year, today.month)[1]
    month_end = today.replace(day=last_day)

    # Generate all dates of the selected month
    all_dates = [month_start + timedelta(days=i) for i in range((month_end - month_start).days + 1)]

    # Identify Sundays
    sundays = [d for d in all_dates if d.weekday() == 6]

    # Fetch active holidays
    holidays = Holiday.objects.filter(
        date__range=(month_start, month_end),
        status='Active'
    ).values_list('date', flat=True)

    # Determine working days (excluding Sundays and holidays)
    working_days = [d for d in all_dates if d not in sundays and d not in holidays]
    total_working_days_in_month = len(working_days)

    # Attendance for selected month
    attendance_records = Attendance.objects.filter(
        employee=currentdata,
        date__range=(month_start, month_end)
    ).order_by("date")

    total_present_days = attendance_records.count()
    total_absent_days = total_working_days_in_month - total_present_days

    # Today's attendance (if viewing current month)
    attendance_data = Attendance.objects.filter(employee=currentdata, date=localdate()).first()

    # Calculate production hours (only for today)
    production_hours = None
    if attendance_data and attendance_data.check_in:
        current_time = now().time()
        check_in = attendance_data.check_in

        current_dt = datetime.combine(localdate(), current_time)
        checkin_dt = datetime.combine(localdate(), check_in)

        duration = current_dt - checkin_dt
        production_hours = round(duration.total_seconds() / 3600, 2)

    context = {
        "attendence": "active",
        "currentdata": currentdata,
        "attendance_data": attendance_data,
        "production_hours": production_hours,
        "total_present_days": total_present_days,
        "total_absent_days": total_absent_days,
        "total_working_days_in_month": total_working_days_in_month,
        "today": today,
        "attendance_records": attendance_records,
        "selected_month": month,
        "selected_year": year,
    }
    return render(request, "attendence.html", context)

@login_required()
def timesheet(request):
    d={'timesheet':"active"}
    return render(request,"timesheet.html",d)



def shift(request):
    allshift=Shift.objects.filter(is_active=True)
    allemp=Schedule.objects.all()
    allshifttype=ShiftType.objects.all()
    print(allshifttype,"sd")
    d={'shift':"active","allemp":allemp,"allshift":allshift,"allshifttype":allshifttype}
    return render(request,"shift.html",d)



def resignation(request):
    d={'resignation':"active"}
    return render(request,"resignation.html",d)



def termination(request):
    d={'termination':"active"}
    return render(request,"termination.html",d)

def salaryslip(request):
    d={'salaryslip':"active"}
    return render(request,"salaryslip.html",d)


def payslip(request):
    d={'payslip':"active"}
    return render(request,"payslip.html",d)

def users(request):
    roles = Role.objects.all()
    permissions = Permission.objects.all()
    context = {
        'roles': roles,
            'permissions': permissions,
        'users': "active"  # highlight menu
    }

    
    return render(request,"users.html",context)

def roles(request):
    d={'roles':"active"}
    return render(request,"roles.html",d)

# def profile(request):
#     d={'profile':"active"}
#     return render(request,"profile.html",d)

@login_required
def profile(request):
    user = request.user

    if request.method == "POST":
        name = request.POST.get("name")
        email = request.POST.get("email")
        password = request.POST.get("password")
        confirm_password = request.POST.get("confirm_password")

        # Update basic details
        if name:
            user.first_name = name

        if email:
            user.email = email

        # Password update (only if entered)
        if password or confirm_password:
            if password != confirm_password:
                messages.error(request, "Passwords do not match")
                return redirect("profile_update")

            user.set_password(password)
            update_session_auth_hash(request, user)

        user.save()
        messages.success(request, "Profile updated successfully")
        return redirect("profile_update")

    return render(request, "profile.html", {
        "user": user,
        "profile": "active"
    })


User = get_user_model()

def ajax_login(request):
    if request.user.is_authenticated:
        return JsonResponse({"success": True, "redirect_url": "/dashboard/"})

    if request.method == "POST":
        email = request.POST.get("username")  # frontend sends email in "username" field
        password = request.POST.get("password")
     

        try:
            user_obj = User.objects.get(email=email)
            user = authenticate(request, username=user_obj.username, password=password)
        except User.DoesNotExist:
            user = None

        if user is not None:
            login(request, user)
            if request.user.is_superuser:
                return JsonResponse({"success": True, "redirect_url": "/"})
            else:
                return JsonResponse({"success": True, "redirect_url": "/emp"})
        else:
            return JsonResponse({"success": False, "error": "Invalid email or password"})

    return JsonResponse({"success": False, "error": "Invalid request"})


def ajax_logout(request):
    if request.user.is_authenticated:
        logout(request)
        # For AJAX: return JSON
        if request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({"success": True, "redirect_url": "/login/"})
        # For normal logout (non-AJAX)
        return redirect("/login/")
    return redirect("/login/")

from datetime import datetime
from decimal import Decimal, InvalidOperation
from django.contrib.auth.hashers import make_password
from django.http import JsonResponse
from django.urls import reverse

from datetime import datetime
from decimal import Decimal, InvalidOperation
from django.contrib.auth.hashers import make_password
from django.contrib.auth.models import User
from django.http import JsonResponse

def to_decimal(value, default=0):
    try:
        return Decimal(value) if value not in (None, "", "None") else Decimal(default)
    except (InvalidOperation, TypeError, ValueError):
        return Decimal(default)

def to_date(value):
    try:
        return datetime.strptime(value, "%Y-%m-%d").date() if value else None
    except ValueError:
        return None

@csrf_exempt
def add_employee(request):
    if request.method == "POST":
        data = request.POST
        files = request.FILES

        try:
            employee_type = data.get("employee_type")
            raw_password = data.get("password")
            if not files.get("aadhar_front") and data.get("employee_type")=="monthly":
                return JsonResponse({"success": False, "message": f"Error: AAdhar front is required"})
            
            if not files.get("aadhar_back") and data.get("employee_type")=="monthly":
                return JsonResponse({"success": False, "message": f"Error: AAdhar back is required"})
            
            if not files.get("aadhar_front") and data.get("employee_type")=="monthly":
                return JsonResponse({"success": False, "message": f"Error: AAdhar front is required"})
            
            if not files.get("pan_image") and data.get("employee_type")=="monthly":
                return JsonResponse({"success": False, "message": f"Error: Pancard is required"})
            
            if not files.get("bank_passbook") and data.get("employee_type")=="monthly":
                return JsonResponse({"success": False, "message": f"Error: Bank passbook is required"})
            
            if not files.get("character_certificate") and data.get("employee_type")=="monthly":
                return JsonResponse({"success": False, "message": f"Error:  character certificate is required"})
            
            



            employee = Employee(
                employee_type=employee_type,
                under=User.objects.get(username='rohan'),
                name=data.get("name"),
                email=data.get("email"),
                phone=data.get("phone"),
                joining_date=to_date(data.get("joining_date")),
                password=make_password(raw_password) if raw_password else None,
               

                # Base Salary
                monthly_salary=to_decimal(data.get("monthly_salary")) if employee_type == "monthly" else None,
                daily_salary=to_decimal(data.get("daily_salary")) if employee_type == "daily" else None,

                # Monthly-only salary components
             
                hra=to_decimal(data.get("hra")) if employee_type == "monthly" else None,
                conveyance=to_decimal(data.get("conveyance")) if employee_type == "monthly" else None,
                special=to_decimal(data.get("special")) if employee_type == "monthly" else None,
                ta=to_decimal(data.get("ta")) if employee_type == "monthly" else None,
                da=to_decimal(data.get("da")) if employee_type == "monthly" else None,
                epfo=to_decimal(data.get("epfo")) if employee_type == "monthly" else None,
                esic=to_decimal(data.get("esic")) if employee_type == "monthly" else None,
                insurance=to_decimal(data.get("insurance")) if employee_type == "monthly" else None,
                professional_tax=to_decimal(data.get("professional_tax")) if employee_type == "monthly" else None,

                # Other fields
                nationality=data.get("nationality"),
                religion=data.get("religion"),
                marital_status=data.get("marital_status"),
                birthday=to_date(data.get("birthday")),
                gender=data.get("gender"),
                pf_fund=data.get("pf_fund"),
                pf_number=data.get("pf_number"),
                esi_number=data.get("esi_number"),
                emergency_number=data.get("emergency_number"),
                bank_name=data.get("bank_name"),
                account_number=data.get("account_number"),
                pan_number=data.get("pan_number"),
                ifsc_number=data.get("ifsc_number"),

                leaves=int(data.get("leaves") or 0),
                bonus=to_decimal(data.get("bonus"), 0),
                labour_tax=to_decimal(data.get("labour_tax"), 0),
                uan_number=data.get("uan_number"),

                department_id=data.get("department"),
                position_id=data.get("position"),
                designation_id=data.get("designation"),
                grade_id=data.get("grade"),
                weekly_off=data.get("weekly_off"),
                description=data.get("description"),
                worker_type=data.get("worker_type"),
                assign_location=data.get("assign_location"),
                role_id=data.get("role"),

                # Profile & Documents
                profile=files.get("profile"),
                aadhar_front=files.get("aadhar_front"),
                aadhar_back=files.get("aadhar_back"),
                pan_image=files.get("pan_image"),
                bank_passbook=files.get("bank_passbook"),
                character_certificate=files.get("character_certificate"),
            )

            employee.save()

            # Create an empty schedule
            Schedule.objects.create(employee=employee)

            # Create Django User if Monthly
            if employee_type == "monthly" and data.get("email") and raw_password:
                if not User.objects.filter(username=data.get("email")).exists():
                   lolo= User.objects.create_user(
                        username=data.get("email"),
                        email=data.get("email"),
                        password=raw_password,
                        first_name=data.get("name"),
                    )
                employee.user=lolo
                employee.save()



            return JsonResponse({
                "success": True,
                "message": "Employee added successfully!",
                "redirect_url": "/employeelist/"
            })

        except Exception as e:
            return JsonResponse({"success": False, "message": f"Error: {str(e)}"})

    return JsonResponse({"success": False, "message": "Invalid request"})


def add_department(request):
    if request.method == "POST":
        name = request.POST.get("name")
        status = request.POST.get("status")

        if not name or status not in ["Active", "Inactive"]:
            return JsonResponse({"success": False, "message": "Please fill all fields correctly."})

        if Department.objects.filter(name=name).exists():
            return JsonResponse({"success": False, "message": "Department already exists."})

        Department.objects.create(name=name, status=status)
        return JsonResponse({"success": True, "message": "Department added successfully!"})

    return JsonResponse({"success": False, "message": "Invalid request"})


def edit_department(request):
    if request.method == "POST":
        dept_id = request.POST.get("id")
        name = request.POST.get("name")
        status = request.POST.get("status")

        try:
            dept = Department.objects.get(id=dept_id)
            dept.name = name
            dept.status = status
            dept.save()
            return JsonResponse({"success": True, "message": "Department updated successfully!"})
        except Department.DoesNotExist:
            return JsonResponse({"success": False, "message": "Department not found."})

    return JsonResponse({"success": False, "message": "Invalid request"})


@csrf_exempt
def delete_department(request):
    if request.method == "POST":
        dept_id = request.POST.get("id")
        try:
            dept = Department.objects.get(id=dept_id)
            dept.delete()
            return JsonResponse({"success": True})
        except Department.DoesNotExist:
            return JsonResponse({"success": False, "message": "Department not found"})
    return JsonResponse({"success": False, "message": "Invalid request"})

def export_departments_pdf(request):
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="departments.pdf"'

    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter

    departments = Department.objects.annotate(employee_count=Count("employee"))
    y = height - 50
    p.setFont("Helvetica-Bold", 14)
    p.drawString(200, y, "Departments Report")

    y -= 40
    p.setFont("Helvetica", 12)
    for dept in departments:
        p.drawString(50, y, f"{dept.name} - {dept.employee_count} Employees - {dept.status}")
        y -= 20
        if y < 50:  # start new page if needed
            p.showPage()
            y = height - 50

    p.showPage()
    p.save()
    return response


def export_departments_excel(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Departments"

    headers = ["Department Name", "No of Employees", "Status"]
    for col_num, header in enumerate(headers, 1):
        sheet[f"{get_column_letter(col_num)}1"] = header

    departments = Department.objects.annotate(employee_count=Count("employee"))
    for row_num, dept in enumerate(departments, 2):
        sheet[f"A{row_num}"] = dept.name
        sheet[f"B{row_num}"] = dept.employee_count
        sheet[f"C{row_num}"] = dept.status

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="departments.xlsx"'
    workbook.save(response)
    return response

   


@login_required
def designation(request):
    # Get all designations with their related department
    data = Designation.objects.select_related("department").all()
    print()

    # Add employee count for each designation (if you linked employees)
    for d in data:
        d.employee_count = d.employees.count() if hasattr(d, 'employees') else 0

    context = {
        "page_active": "designation",  # ✅ used in sidebar to set active class
        "designations": data, 
                 "department"    :Department.objects.all()     # ✅ actual queryset for template loop
    }
    return render(request, "designation.html", context)


@login_required
def add_designation(request):
    if request.method == "POST":
        name = request.POST.get("name")
        dept_id = request.POST.get("department")   # get department id from form
        status = request.POST.get("status")
        print(dept_id)

        if not name or not dept_id or not status:
            return JsonResponse({"success": False, "error": "All fields are required."})

        try:
            department = Department.objects.get(id=dept_id)
        except Department.DoesNotExist:
            return JsonResponse({"success": False, "error": "Invalid department selected."})

        designation = Designation.objects.create(
            name=name,
            department=department,
            status=status
        )

        return JsonResponse({"success": True, "id": designation.id})

    return JsonResponse({"success": False, "error": "Invalid request"})


@login_required
def edit_designation(request):
    if request.method == "POST":
        designation_id = request.POST.get("id")
        name = request.POST.get("name")
        department_id = request.POST.get("department")
        status = request.POST.get("status")

        if not designation_id or not name or not department_id or not status:
            return JsonResponse({"success": False, "error": "All fields are required."})

        try:
            designation = Designation.objects.get(id=designation_id)
            department = Department.objects.get(id=department_id)

            designation.name = name
            designation.department = department
            designation.status = status
            designation.save()

            return JsonResponse({"success": True})
        except Designation.DoesNotExist:
            return JsonResponse({"success": False, "error": "Designation not found"})
        except Department.DoesNotExist:
            return JsonResponse({"success": False, "error": "Invalid department"})
    return JsonResponse({"success": False, "error": "Invalid request"})

@login_required
def delete_designation(request, pk):
    designation = get_object_or_404(Designation, pk=pk)
    designation.delete()
    return JsonResponse({"success": True})

# Export Excel
@login_required
def export_designations_excel(request):
    data = Designation.objects.all().values("name", "department__name", "status")
    df = pd.DataFrame(list(data))
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Designations")
    buffer.seek(0)
    response = HttpResponse(buffer, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="designations.xlsx"'
    return response

# Export PDF
@login_required
def export_designations_pdf(request):
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="designations.pdf"'
    p = canvas.Canvas(response)
    y = 800
    p.setFont("Helvetica", 12)
    p.drawString(50, y, "Designation | Department | Status")
    for designation in Designation.objects.select_related("department").all():
        y -= 20
        line = f"{designation.name} | {designation.department.name} | {designation.status}"
        p.drawString(50, y, line)
    p.showPage()
    p.save()
    return response


@login_required
def grade(request):
    departments = Grade.objects.annotate(employee_count=Count('employee'))
    context = {
        "grade": "active",
        "grades": departments,
    }
    return render(request, "grade.html", context)


@login_required
def shiftlist(request):
    departments = Shift.objects.all()
    allshifttype=ShiftType.objects.all()

    context = {
        "listshift": "active",
        "departments": departments,
        "allshifttype":allshifttype,
    }
    return render(request, "createshift.html", context)

def add_grade(request):
    if request.method == "POST":
        print("hello")
        name = request.POST.get("name")
        status = request.POST.get("status")

        if not name or status not in ["Active", "Inactive"]:
            return JsonResponse({"success": False, "message": "Please fill all fields correctly."})

        if Grade.objects.filter(name=name).exists():
            return JsonResponse({"success": False, "message": "Grade already exists."})

        Grade.objects.create(name=name, status=status)
        return JsonResponse({"success": True, "message": "Grade added successfully!"})

    return JsonResponse({"success": False, "message": "Invalid request"})


def edit_grade(request):
    if request.method == "POST":
        dept_id = request.POST.get("id")
        name = request.POST.get("name")
        status = request.POST.get("status")

        try:
            dept = Grade.objects.get(id=dept_id)
            dept.name = name
            dept.status = status
            dept.save()
            return JsonResponse({"success": True, "message": "Grade updated successfully!"})
        except Grade.DoesNotExist:
            return JsonResponse({"success": False, "message": "Grade not found."})

    return JsonResponse({"success": False, "message": "Invalid request"})


@csrf_exempt
def delete_grade(request):
    if request.method == "POST":
        dept_id = request.POST.get("id")
        try:
            dept = Grade.objects.get(id=dept_id)
            dept.delete()
            return JsonResponse({"success": True})
        except Grade.DoesNotExist:
            return JsonResponse({"success": False, "message": "Grade not found"})
    return JsonResponse({"success": False, "message": "Invalid request"})


def export_grades_pdf(request):
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="grades.pdf"'

    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter

    departments = Grade.objects.annotate(employee_count=Count("employee"))
    y = height - 50
    p.setFont("Helvetica-Bold", 14)
    p.drawString(200, y, "Grades Report")

    y -= 40
    p.setFont("Helvetica", 12)
    for dept in departments:
        p.drawString(50, y, f"{dept.name} - {dept.employee_count} Employees - {dept.status}")
        y -= 20
        if y < 50:  # start new page if needed
            p.showPage()
            y = height - 50

    p.showPage()
    p.save()
    return response


def export_grades_excel(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Grades"

    headers = ["Grades Name", "No of Employees", "Status"]
    for col_num, header in enumerate(headers, 1):
        sheet[f"{get_column_letter(col_num)}1"] = header

    departments = Grade.objects.annotate(employee_count=Count("employee"))
    for row_num, dept in enumerate(departments, 2):
        sheet[f"A{row_num}"] = dept.name
        sheet[f"B{row_num}"] = dept.employee_count
        sheet[f"C{row_num}"] = dept.status

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="Grades.xlsx"'
    workbook.save(response)
    return response





@login_required
def position(request):
    departments = Position.objects.annotate(employee_count=Count('employee'))
    context = {
        "position": "active",
        "position": departments,
    }
    return render(request, "position.html", context)


@login_required
def shifttype(request):
    departments = ShiftType.objects.all()
    context = {
        "shifttype": "active",
        "position": departments,
    }
    return render(request, "shifttype.html", context)


def add_position(request):
    if request.method == "POST":


        name = request.POST.get("name")
        status = request.POST.get("status")

        if not name or status not in ["Active", "Inactive"]:
            return JsonResponse({"success": False, "message": "Please fill all fields correctly."})

        if Position.objects.filter(name=name).exists():
            return JsonResponse({"success": False, "message": "Position already exists."})

        Position.objects.create(name=name, status=status)
        return JsonResponse({"success": True, "message": "Position added successfully!"})

    return JsonResponse({"success": False, "message": "Invalid request"})



def add_shifttype(request):
    if request.method == "POST":


        name = request.POST.get("name")
        status = request.POST.get("status")

        if not name or status not in ["Active", "Inactive"]:
            return JsonResponse({"success": False, "message": "Please fill all fields correctly."})

        if Position.objects.filter(name=name).exists():
            return JsonResponse({"success": False, "message": "Position already exists."})

        ShiftType.objects.create(name=name, status=status)
        return JsonResponse({"success": True, "message": "Position added successfully!"})

    return JsonResponse({"success": False, "message": "Invalid request"})



def edit_position(request):
    if request.method == "POST":
        dept_id = request.POST.get("id")
        name = request.POST.get("name")
        status = request.POST.get("status")

        try:
            dept = Position.objects.get(id=dept_id)
            dept.name = name
            dept.status = status
            dept.save()
            return JsonResponse({"success": True, "message": "Position updated successfully!"})
        except Position.DoesNotExist:
            return JsonResponse({"success": False, "message": "Position not found."})

    return JsonResponse({"success": False, "message": "Invalid request"})




def edit_shifttype(request):
    if request.method == "POST":
        dept_id = request.POST.get("id")
        name = request.POST.get("name")
        status = request.POST.get("status")

        try:
            dept = ShiftType.objects.get(id=dept_id)
            dept.name = name
            dept.status = status
            dept.save()
            return JsonResponse({"success": True, "message": "Position updated successfully!"})
        except Position.DoesNotExist:
            return JsonResponse({"success": False, "message": "Position not found."})

    return JsonResponse({"success": False, "message": "Invalid request"})

@csrf_exempt
def delete_position(request):
    if request.method == "POST":
        dept_id = request.POST.get("id")
        try:
            dept = Position.objects.get(id=dept_id)
            dept.delete()
            return JsonResponse({"success": True})
        except Position.DoesNotExist:
            return JsonResponse({"success": False, "message": "Position not found"})
    return JsonResponse({"success": False, "message": "Invalid request"})


@csrf_exempt
def delete_shifttype(request):
    if request.method == "POST":
        dept_id = request.POST.get("id")
        try:
            dept = ShiftType.objects.get(id=dept_id)
            dept.delete()
            return JsonResponse({"success": True})
        except Position.DoesNotExist:
            return JsonResponse({"success": False, "message": "Position not found"})
    return JsonResponse({"success": False, "message": "Invalid request"})


def export_position_pdf(request):
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="Position.pdf"'

    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter

    departments = Position.objects.annotate(employee_count=Count("employee"))
    y = height - 50
    p.setFont("Helvetica-Bold", 14)
    p.drawString(200, y, "Position Report")

    y -= 40
    p.setFont("Helvetica", 12)
    for dept in departments:
        p.drawString(50, y, f"{dept.name} - {dept.employee_count} Employees - {dept.status}")
        y -= 20
        if y < 50:  # start new page if needed
            p.showPage()
            y = height - 50

    p.showPage()
    p.save()
    return response


def export_position_excel(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Position"

    headers = ["Position Name", "No of Employees", "Status"]
    for col_num, header in enumerate(headers, 1):
        sheet[f"{get_column_letter(col_num)}1"] = header

    departments = Position.objects.annotate(employee_count=Count("employee"))
    for row_num, dept in enumerate(departments, 2):
        sheet[f"A{row_num}"] = dept.name
        sheet[f"B{row_num}"] = dept.employee_count
        sheet[f"C{row_num}"] = dept.status

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="Position.xlsx"'
    workbook.save(response)
    return response

# @csrf_exempt
# def add_role(request):
#     if request.method == "POST":
#         name = request.POST.get("role_name")
#         permissions = request.POST.getlist("permissions[]")

#         # create the role
#         role = Role.objects.create(name=name)

#         # assign unique permissions
#         for perm_name in set(permissions):  # set() prevents duplicates in input
#             perm, _ = Permission.objects.get_or_create(name=perm_name)
#             RolePermission.objects.get_or_create(role=role, permission=perm)

#         return JsonResponse({"success": True})

#     return JsonResponse({"success": False, "error": "Invalid request"})


# @csrf_exempt
# def add_role(request):
#     if request.method == "POST":
#         name = request.POST.get("role_name")
#         permissions = request.POST.getlist("permissions[]")

#         # create the role
#         #role = Role.objects.create(name=name)

#         print(permissions,type(permissions))
     


#         # create a separate permission record for this role
#         for perm_name in permissions:  # set() prevents duplicates from request
#             try:
#                 # perm = Permission.objects.create(name=perm_name)
#                 print("")
#             except:
#                 pass
  

       
#             # for i in perm_name:
#             #     print(i)
#             # RolePermission.objects.create(role=role, permission=perm)

#         return JsonResponse({"success": True})

#     return JsonResponse({"success": False, "error": "Invalid request"})


@csrf_exempt
def add_role(request):
    if request.method == "POST":
        name = request.POST.get("role_name")
        permissions_raw = request.POST.getlist("permissions[]")

        # If only one item but it's a comma-separated string, split it
        if len(permissions_raw) == 1 and "," in permissions_raw[0]:
            permissions = [p.strip() for p in permissions_raw[0].split(",") if p.strip()]
        else:
            permissions = permissions_raw

        # create the role
        role = Role.objects.create(name=name)

        # create a separate permission record for this role
        for perm_name in set(permissions):
            # try:
            #     perm = Permission.objects.create(name=perm_name)
            #     RolePermission.objects.create(role=role, permission=perm)
            # except:

            #     pass
                perm, _ = Permission.objects.get_or_create(name=perm_name)
                RolePermission.objects.get_or_create(role=role, permission=perm)
            

        return JsonResponse({"success": True})

    return JsonResponse({"success": False, "error": "Invalid request"})


@csrf_exempt
def edit_role(request):
    if request.method == "POST":
        role_id = request.POST.get("role_id")
        if not role_id:
            return JsonResponse({"success": False, "error": "No role ID provided"})

        try:
            role = Role.objects.get(id=int(role_id))
        except (Role.DoesNotExist, ValueError):
            return JsonResponse({"success": False, "error": "Invalid role ID"})

        # update role name
        name = request.POST.get("name")
        role.name = name
        role.save()

        # handle permissions
        new_permissions = set(request.POST.getlist("permissions"))
        current_permissions = set(role.role_permissions.values_list("permission__name", flat=True))

        # add new permissions
        for perm_name in new_permissions - current_permissions:
            perm, _ = Permission.objects.get_or_create(name=perm_name)
            RolePermission.objects.get_or_create(role=role, permission=perm)

        # remove unchecked permissions
        RolePermission.objects.filter(
            role=role,
            permission__name__in=(current_permissions - new_permissions)
        ).delete()

        return JsonResponse({"success": True})

    return JsonResponse({"success": False, "error": "Invalid request"})


@csrf_exempt
def delete_role(request):
    if request.method == "POST":
        role_id = request.POST.get("role_id")
        try:
            role = Role.objects.get(id=role_id)
            role.delete()  
            return JsonResponse({"success": True})
        except Role.DoesNotExist:
            return JsonResponse({"success": False, "error": "Role not found"})
    return JsonResponse({"success": False, "error": "Invalid request"})



# def export_employee_pdf(request):
#     response = HttpResponse(content_type="application/pdf")
#     response["Content-Disposition"] = 'attachment; filename="Employee_List.pdf"'

#     p = canvas.Canvas(response, pagesize=letter)
#     width, height = letter
#     y = height - 40

#     p.setFont("Helvetica-Bold", 14)
#     p.drawString(200, y, "Employee List")

#     y -= 30
#     p.setFont("Helvetica-Bold", 10)
#     headers = ["Emp ID", "Name", "Email", "Phone", "Designation", "Joining Date", "Status"]
#     x_positions = [30, 80, 180, 300, 380, 460, 540]
#     for i, header in enumerate(headers):
#         p.drawString(x_positions[i], y, header)

#     y -= 20
#     p.setFont("Helvetica", 9)

#     employees = Employee.objects.select_related("designation").all()
#     for emp in employees:
#         row = [
#             emp.id,
#             emp.name,
#             emp.email or "",
#             emp.phone,
#             emp.designation.name if emp.designation else "",
#             emp.joining_date.strftime("%Y-%m-%d") if emp.joining_date else "",
#             emp.status,
#         ]
#         for i, item in enumerate(row):
#             p.drawString(x_positions[i], y, str(item))
#         y -= 15
#         if y < 40:
#             p.showPage()
#             y = height - 40
#             p.setFont("Helvetica", 9)

#     p.showPage()
#     p.save()
#     return response

from django.http import HttpResponse
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors

def export_employee_pdf(request):
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="Employee_List.pdf"'

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        rightMargin=30,
        leftMargin=30,
        topMargin=40,
        bottomMargin=30
    )

    styles = getSampleStyleSheet()
    elements = []

    # Title (acts like <h2>)
    elements.append(
        Paragraph("<b>Employee List</b>", styles["Title"])
    )
    elements.append(Paragraph("<br/>", styles["Normal"]))

    # Table Header (HTML-style)
    table_data = [[
        "Emp ID", "Name", "Email", "Phone",
        "Designation", "Joining Date", "Status"
    ]]

    employees = Employee.objects.select_related("designation").all()

    for emp in employees:
        table_data.append([
            emp.id,
            emp.name,
            emp.email or "",
            emp.phone or "",
            emp.designation.name if emp.designation else "",
            emp.joining_date.strftime("%Y-%m-%d") if emp.joining_date else "",
            emp.status,
        ])

    table = Table(
        table_data,
        colWidths=[60, 120, 180, 100, 120, 100, 80]
    )

    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONT", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("ALIGN", (0, 1), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
        ("TOPPADDING", (0, 0), (-1, 0), 10),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
    ]))

    elements.append(table)

    doc.build(elements)
    return response


def export_employee_excel(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Employees"

    headers = ["Emp ID", "Name", "Email", "Phone", "Designation", "Joining Date", "Status"]
    for col_num, header in enumerate(headers, 1):
        sheet[f"{get_column_letter(col_num)}1"] = header

    employees = Employee.objects.select_related("designation").all()
    for row_num, emp in enumerate(employees, 2):
        data = [
            emp.id,
            emp.name,
            emp.email or "",
            emp.phone,
            emp.designation.name if emp.designation else "",
            emp.joining_date.strftime("%Y-%m-%d") if emp.joining_date else "",
            emp.status,
        ]
        for col_num, value in enumerate(data, 1):
            sheet[f"{get_column_letter(col_num)}{row_num}"] = value

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="Employee_List.xlsx"'
    workbook.save(response)
    return response



def delete_employee(request):
	if request.method == "POST":
		emp_id = request.POST.get("employee_id")
		employee = get_object_or_404(Employee, id=emp_id)
		
		# Get employee email before deleting
		emp_email = employee.email
		
		# Delete the employee record
		employee.delete()

		# Try to delete the related User with the same email
		User.objects.filter(email=emp_email).delete()
	
	return redirect(request.META.get('HTTP_REFERER', 'employee_list'))



@csrf_exempt  # Only if you're not using {% csrf_token %}, else REMOVE this
def add_holiday_ajax(request):
    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        date_str = request.POST.get('date', '').strip()
        description = request.POST.get('description', '').strip()
        status = request.POST.get('status', 'Active')
        print(date_str)

        # Basic validation
        if not title or not date_str:
            return JsonResponse({'success': False, 'error': 'Title and Date are required.'})

        try:
            # Parse format: DD-MM-YYYY
            date = datetime.strptime(date_str, "%d-%m-%Y").date()
        except ValueError:
            return JsonResponse({'success': False, 'error': 'Invalid date format. Use DD-MM-YYYY.'})

        # Save the Holiday
        Holiday.objects.create(
            title=title,
            date=date,
            description=description,
            status=status
        )

        return JsonResponse({'success': True})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})



@csrf_exempt
def update_holiday_ajax(request):
    if request.method == 'POST':
        id = request.POST.get('id')
        title = request.POST.get('title')
        date_str = request.POST.get('date')
        description = request.POST.get('description')
        status = request.POST.get('status')

        try:
            holiday = Holiday.objects.get(id=id)
        except Holiday.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Holiday not found'})

        try:
            date = datetime.strptime(date_str, '%d-%m-%Y').date()
        except ValueError:
            return JsonResponse({'success': False, 'error': 'Invalid date format (dd-mm-yyyy)'})

        holiday.title = title
        holiday.date = date
        holiday.description = description
        holiday.status = status
        holiday.save()

        return JsonResponse({'success': True})

    return JsonResponse({'success': False, 'error': 'Invalid method'})


@login_required
@csrf_exempt  # Optional if you use CSRF token in AJAX
def delete_holiday(request):
    print("hello")
    if request.method == 'POST':
        print(request.POST.get('id'),"lol")
        holiday_id = request.POST.get('id')
        try:
            holiday = Holiday.objects.get(id=holiday_id)
            holiday.delete()
            return JsonResponse({'status': 'success'})
        except Holiday.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Not found'})
    return JsonResponse({'status': 'error', 'message': 'Invalid request'})



@login_required
@csrf_exempt
def add_leave(request):
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        description = request.POST.get("description", "").strip()

        if not name:
            return JsonResponse({"status": "error", "message": "Leave name is required."})

        Leave.objects.create(name=name, description=description)
        return JsonResponse({"status": "success"})

    return JsonResponse({"status": "error", "message": "Invalid request method"})



@csrf_exempt
@login_required
def delete_leave(request, leave_id):
    if request.method == "POST":
        try:
            leave = Leave.objects.get(id=leave_id)
            leave.delete()
            return JsonResponse({'success': True})
        except Leave.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Leave not found'})
    return JsonResponse({'success': False, 'error': 'Invalid request'})

import json
@csrf_exempt
@login_required
def toggle_leave_status(request, leave_id):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            status = data.get("status", "Inactive")
            leave = Leave.objects.get(id=leave_id)
            leave.status = status
            leave.save()
            return JsonResponse({'success': True})
        except Leave.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Leave not found'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Invalid request'})


@csrf_exempt
@login_required
def add_shift(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        start_time = request.POST.get('start_time')
        end_time = request.POST.get('end_time')
        break_duration = request.POST.get('break_duration')  # Format: "HH:MM:SS"
        shift_type = request.POST.get('shift_type')
        

        try:
            shift = Shift.objects.create(
                name=name,
                start_time=start_time,
                end_time=end_time,
                break_duration=parse_duration(break_duration),
                shift_type=shift_type,
     
            )
            return JsonResponse({'status': 'success'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
from datetime import timedelta

def parse_duration(time_str):
    try:
        h, m = time_str.split(':')
        return timedelta(hours=int(h), minutes=int(m))
    except:
        return timedelta(0)

def update_shift(request):
    if request.method == 'POST':
        try:
            shift = Shift.objects.get(id=request.POST.get("id"))
            shift.name = request.POST.get("name")
            shift.start_time = request.POST.get("start_time")
            shift.end_time = request.POST.get("end_time")

            break_str = request.POST.get("break_duration") or "00:00"
            shift.break_duration = parse_duration(break_str)  # ✅ convert to timedelta

            shift.shift_type = request.POST.get("shift_type")
            shift.is_active = request.POST.get("is_active") == 'true'

            shift.save()
            return JsonResponse({'success': True})
        except Shift.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Shift not found'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})


@csrf_exempt  # Optional if you're not using {% csrf_token %}
def delete_shift(request):
    if request.method == 'POST':
        try:
            shift_id = request.POST.get('id')
            shift = Shift.objects.get(id=shift_id)
            shift.delete()
            return JsonResponse({'success': True})
        except Shift.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Shift not found'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
 
@csrf_exempt
def add_schedule(request):
    if request.method == "POST":
        try:
            employee_id = request.POST.get("employee_id")
            shift_id = request.POST.get("shift_id")

            if not employee_id or not shift_id:
                return JsonResponse({"success": False, "error": "All fields are required."})

            try:
                employee = Employee.objects.get(id=employee_id)
            except Employee.DoesNotExist:
                return JsonResponse({"success": False, "error": "Employee not found."})

            try:
                shift = Shift.objects.get(id=shift_id)
            except Shift.DoesNotExist:
                return JsonResponse({"success": False, "error": "Shift not found."})

            # Check for existing schedule and update or create accordingly
            schedule = Schedule.objects.filter(employee=employee).first()

            if schedule:
                schedule.shift = shift
                schedule.save()
                return JsonResponse({"success": True, "message": "Schedule updated."})
            else:
                Schedule.objects.create(employee=employee, shift=shift)
                return JsonResponse({"success": True, "message": "Schedule created."})

        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})

    return JsonResponse({"success": False, "error": "Invalid request method."})



@login_required
def addattendence(request):
    # departments = Grade.objects.annotate(employee_count=Count('employee'))
    # allemp=Employee.objects.all()
    if request.user.is_superuser:
        allemp = Employee.objects.all()
    else:
        allemp = Employee.objects.filter(under=request.user).exclude(user=request.user)
    context = {
        "addattendence": "active",
        "allemp":allemp,
        # "grades": departments,
    }
    return render(request, "addattendence.html", context)


def get_employee_shift(request):
    employee_id = request.GET.get("employee_id")
    try:
        schedule = Schedule.objects.select_related("shift").get(employee_id=employee_id)
        shift = schedule.shift
        return JsonResponse({
            "success": True,
            "shift": {
                "name": shift.name,
                "start_time": shift.start_time.strftime("%H:%M"),
                "end_time": shift.end_time.strftime("%H:%M"),
                "break_duration": str(shift.break_duration),
                "type": shift.shift_type,
            }
        })
    except Schedule.DoesNotExist:
        return JsonResponse({"success": False, "error": "No shift assigned."})
    

def get_attendance_shift(request):
    emp_id = request.GET.get("employee_id")
    date_str = request.GET.get("date")  # Format: YYYY-MM-DD

  
    try:
        date = datetime.strptime(date_str, "%Y-%m-%d").date()
    except:
    
        return JsonResponse({"success": False, "error": "Invalid date format."})

    data = {}
    try:
        schedule = Schedule.objects.select_related('shift').get(employee_id=emp_id)
        shift = schedule.shift
        data["shift"] = {
            "start_time": shift.start_time.strftime("%H:%M"),
            "end_time": shift.end_time.strftime("%H:%M"),
        }
    except Schedule.DoesNotExist:
        return JsonResponse({"success": False, "error": "No shift assigned."})

    try:
        attendance = Attendance.objects.get(employee_id=emp_id, date=date)
        data["attendance"] = {
            "check_in": attendance.check_in.strftime("%H:%M") if attendance.check_in else "",
            "check_out": attendance.check_out.strftime("%H:%M") if attendance.check_out else "",
            "remarks": attendance.remarks or "",
            "type": attendance.type or "",
        }
        data["already_exists"] = True
    except Attendance.DoesNotExist:
        data["attendance"] = {}
        data["already_exists"] = False

    return JsonResponse({"success": True, **data})

from django.utils.timezone import now

def save_attendance(request):
    if request.method == "POST":
        emp_id = request.POST.get("employee_id")
        date_str = request.POST.get("date")
        check_in = request.POST.get("check_in")
        check_out = request.POST.get("check_out")
        remarks = request.POST.get("remarks", "")
       
        # Parse date and check-in
        date = datetime.strptime(date_str, "%Y-%m-%d").date()
        check_in_time = datetime.strptime(check_in, "%H:%M").time()

        check_out_time = None
        if check_out:  # ✅ Only parse if not empty
            try:
                check_out_time = datetime.strptime(check_out, "%H:%M").time()
            except ValueError:
                return JsonResponse({"success": False, "error": "Invalid check-out format."})

        # Get shift details
        try:
            shift = Schedule.objects.get(employee_id=emp_id).shift
        except Schedule.DoesNotExist:
            return JsonResponse({"success": False, "error": "No shift assigned."})

        shift_start = shift.start_time
        shift_end = shift.end_time
        break_duration = shift.break_duration
        shift_total = shift.total_work_duration()

        # Determine work duration if check_out is provided
        work_duration = timedelta()
        if check_out_time:
            in_dt = datetime.combine(date, check_in_time)
            out_dt = datetime.combine(date, check_out_time)
            if out_dt < in_dt:
                out_dt += timedelta(days=1)
            work_duration = out_dt - in_dt

        # Determine attendance type
        if datetime.combine(date, check_in_time) > datetime.combine(date, shift_start) + timedelta(minutes=15):
            attendance_type = "Late"
        elif check_out_time and work_duration < timedelta(hours=4):
            attendance_type = "Half Day"
        elif check_out_time:
            attendance_type = "Full Day"
        else:
            attendance_type = "Late"  # Default if only check-in

        # Save attendance
        attendance, created = Attendance.objects.get_or_create(employee_id=emp_id, date=date)
        attendance.check_in = check_in_time
        attendance.check_out = check_out_time  # Can be None
        attendance.remarks = remarks
        attendance.type = attendance_type
        attendance.save()

        return JsonResponse({"success": True, "msg": "Attendance saved."})


@csrf_exempt
def set_under(request):
    if request.method == "POST":
   
        emp_id = request.POST.get("employee_id")
        under_id = request.POST.get("under_id")

        if under_id =="admin":
            employee = Employee.objects.get(id=emp_id)
            under_user = User.objects.get(username="rohan")
            employee.under = under_user
            employee.save()
            return JsonResponse({"success": True})
    


        try:
         
 
            employee = Employee.objects.get(id=emp_id)
            employeee = Employee.objects.get(id=under_id)
            under_user = User.objects.get(email=employeee.email)
            employee.under = under_user
         
            employee.save()
            return JsonResponse({"success": True})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})
        
@csrf_exempt  # Required if CSRF token is not passed properly
def submit_leave(request):
  
    if request.method == "POST":
        try:
            employee = Employee.objects.get(user=request.user)  # or get from POST if admin
            leave_type = request.POST.get("leave_type")
            duration_type = request.POST.get("duration_type")
            from_date = request.POST.get("from_date")
            to_date = request.POST.get("to_date")
            half_day_date = request.POST.get("half_day_date")
            half_type = request.POST.get("half_type")
            reason = request.POST.get("reason")
            attachment = request.FILES.get("attachment")

            

            # Parse dates
            from_date = datetime.strptime(from_date, "%d-%m-%Y").date() if from_date else None
            
            to_date = datetime.strptime(to_date, "%d-%m-%Y").date() if to_date else None
        
            half_day_date = datetime.strptime(half_day_date, "%d-%m-%Y").date() if half_day_date else None
            kjjkfdkjd=Leave.objects.get(name=leave_type)


            abc=LeaveApplication.objects.create(
                employee=employee,
                leave_type=kjjkfdkjd,
                duration_type=duration_type,
                from_date=from_date,
                to_date=to_date,
                half_day_date=half_day_date,
                half_type=half_type,
                reason=reason,
                attachment=attachment,
            )

            return JsonResponse({"success": True})
        except Exception as e:
            print(e)
            return JsonResponse({"success": False, "error": str(e)})

    return JsonResponse({"success": False, "error": "Invalid request method"})
from django.utils import timezone


# @csrf_exempt
# @login_required
# def update_leave_status(request):
#     if request.method == "POST":
#         leave_id = request.POST.get("leave_id")
#         try:
#             leave_app = LeaveApplication.objects.get(id=leave_id)
#         except LeaveApplication.DoesNotExist:
#             return JsonResponse({"success": False, "message": "Leave not found."})

#         if request.user.is_superuser:
#             leave_app.superadmin_approval = True
#             leave_app.superadmin_approved_on = timezone.now().date()
#             leave_app.save()

#             # Create attendance records for approved leave
#             if leave_app.duration_type == 'full' and leave_app.from_date and leave_app.to_date:
#                 current_date = leave_app.from_date
#                 while current_date <= leave_app.to_date:
#                     Attendance.objects.update_or_create(
#                         employee=leave_app.employee,
#                         date=current_date,
#                         defaults={
#                             "type": "Absent",
#                             "remarks": "On Leave (Approved by Superadmin)"
#                         }
#                     )
#                     current_date += timedelta(days=1)

#             elif leave_app.duration_type == 'half' and leave_app.half_day_date:
#                 Attendance.objects.update_or_create(
#                     employee=leave_app.employee,
#                     date=leave_app.half_day_date,
#                     defaults={
#                         "type": "Half Day",
#                         "remarks": f"Half Day Leave - {leave_app.half_type.title()} Half (Approved by Superadmin)"
#                     }
#                 )

#             return JsonResponse({"success": True})

#         elif leave_app.employee.under == request.user:
#             leave_app.admin_approval = True
#             leave_app.admin_approved_on = timezone.now().date()
#             leave_app.save()
#             return JsonResponse({"success": True})

#         else:
#             return JsonResponse({"success": False, "message": "Permission denied."})

#     return JsonResponse({"success": False, "message": "Invalid request."})



@csrf_exempt
@login_required
def update_leave_status(request):
    if request.method == "POST":
        leave_id = request.POST.get("leave_id")
        action = request.POST.get("action")

        try:
            leave_app = LeaveApplication.objects.get(id=leave_id)
        except LeaveApplication.DoesNotExist:
            return JsonResponse({"success": False, "message": "Leave not found."})

        user = request.user

        # --- Approve logic ---
        if action == "approve":
            if user.is_superuser:
                leave_app.superadmin_approval = True
                leave_app.superadmin_approved_on = timezone.now().date()
                leave_app.save()

                # Add attendance records
                if leave_app.duration_type == 'full':
                    current_date = leave_app.from_date
                    while current_date <= leave_app.to_date:
                        Attendance.objects.update_or_create(
                            employee=leave_app.employee,
                            date=current_date,
                            defaults={
                                "type": "Absent",
                                "remarks": "On Leave (Approved by Superadmin)"
                            }
                        )
                        current_date += timedelta(days=1)
                elif leave_app.duration_type == 'half' and leave_app.half_day_date:
                    Attendance.objects.update_or_create(
                        employee=leave_app.employee,
                        date=leave_app.half_day_date,
                        defaults={
                            "type": "Half Day",
                            "remarks": f"Half Day Leave - {leave_app.half_type.title()} Half (Approved by Superadmin)"
                        }
                    )

                return JsonResponse({"success": True})

            elif leave_app.employee.under == user:
                leave_app.admin_approval = True
                leave_app.admin_approved_on = timezone.now().date()
                leave_app.save()
                return JsonResponse({"success": True,"message": "Cancelled."})

            return JsonResponse({"success": False, "message": "Permission denied."})

        # --- Decline logic ---
        elif action == "decline":
       
            if user.is_superuser:
           
                leave_app.superadmin_approval = False
                leave_app.superadmin_cancelled = True
                leave_app.superadmin_approved_on = None
                leave_app.save()
             
                return JsonResponse({"success": True})

            elif leave_app.employee.under == user:
                leave_app.admin_approval = False
                leave_app.admin_cancelled = True
                leave_app.admin_approved_on = None
                leave_app.save()
                return JsonResponse({"success": True})

            return JsonResponse({"success": False, "message": "Permission denied."})

@csrf_exempt
@login_required
def update_advance_salary_status(request):
    if request.method == "POST":
        salary_id = request.POST.get("salary_id")
        action = request.POST.get("action")

        # NEW fields coming from JS
        interest_applicable = request.POST.get("interest_applicable") == "true"
        interest_rate = request.POST.get("interest_rate")

        try:
            app = AdvanceSalaryApplication.objects.get(id=salary_id)
        except AdvanceSalaryApplication.DoesNotExist:
            return JsonResponse({"success": False, "message": "Application not found."})

        user = request.user

        # ================================
        # APPROVE LOGIC
        # ================================
        if action == "approve":

            # Superadmin Approval
            if user.is_superuser:
                app.superadmin_approval = True
                app.superadmin_approved_on = timezone.now().date()

                # ---- NEW: Save interest ----
                if interest_applicable:
                    app.interest_applicable = True
                    app.interest_rate = interest_rate
                else:
                    app.interest_applicable = False
                    app.interest_rate = None

                app.save()
                return JsonResponse({"success": True})

            # Admin Approval
            elif app.employee.under == user:
                app.admin_approval = True
                app.admin_approved_on = timezone.now().date()
                app.save()
                return JsonResponse({"success": True})

            return JsonResponse({"success": False, "message": "Permission denied."})

        # ================================
        # DECLINE LOGIC
        # ================================
        elif action == "decline":

            # Superadmin Decline
            if user.is_superuser:
                app.superadmin_cancelled = True
                app.superadmin_approval = False
                app.superadmin_approved_on = None
                app.save()
                return JsonResponse({"success": True})

            # Admin Decline
            elif app.employee.under == user:
                app.admin_cancelled = True
                app.admin_approval = False
                app.admin_approved_on = None
                app.save()
                return JsonResponse({"success": True})

            return JsonResponse({"success": False, "message": "Permission denied."})

        return JsonResponse({"success": False, "message": "Invalid request."})
    
@login_required()
def salarylist(request):
    today = datetime.today()
    current_month = today.month
    current_year = today.year
    if request.user.is_superuser:
        employees = Employee.objects.select_related(
            "designation", "role", "department", "grade", "position"
        ).all().order_by("-id")
    else:
        employees = Employee.objects.select_related(
            "designation", "role", "department", "grade", "position"
        ).filter(under=request.user).exclude(user=request.user).order_by("-id")

    context = {
        "salarylist": "active",
        "employee": employees,
        "current_month": current_month,
        "current_year": current_year,
    }
    return render(request, "salarylist.html", context)

# @login_required()
# def geneartedsalaryslip(request,data):
#     d={'salaryslip':"active","data": MonthlySalary.objects.filter(employee__id=data)}
#     return render(request,"geneartedsalarylist.html",d)




from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from calendar import monthrange

@login_required()
def geneartedsalaryslip(request, data):
    salary_list = MonthlySalary.objects.filter(employee__id=data)
    today=datetime.today()
    emp=None

    for salary_obj in salary_list:
        emp = salary_obj.employee

        # attendance data (example – adjust if already available)
        year = salary_obj.year
        month = salary_obj.month
        total_days_in_month = monthrange(year, month)[1]
        total_present = salary_obj.present_days
        present_days = getattr(salary_obj, 'present_days', 27)
        holidayss = len(Holiday.objects.filter(
            date__month=today.month,
            date__year=today.year,
            status="Active"
        ).values_list("date", flat=True))
        
   
    if emp is None:
        return redirect('/salarylist')
    
  
    
    if emp.joining_date:
 
        joining_date = emp.joining_date

            # Check if joined in same month & year
        
        if joining_date.year == today.year and joining_date.month == today.month:
            
                join_day = joining_date.day

                if join_day <= 7:
                    holidays = 4
                elif join_day <= 14:
                    holidays = 3
                elif join_day <= 21:
                    holidays = 2
                else:
                    holidays = 1
        else:
                
                holidays = getattr(salary_obj, 'holidays', 4)
        holidays=holidays+holidayss




        total_present = present_days + holidays

        def prorated_amount(monthly_salary, percentage):
            monthly_amount = (monthly_salary * percentage) / 100
            per_day_amount = monthly_amount / total_days_in_month
            return round(per_day_amount * total_present, 2)

        # Allowances
        conveyance = prorated_amount(emp.monthly_salary, emp.conveyance)
        hra        = prorated_amount(emp.monthly_salary, emp.hra)
        special    = prorated_amount(emp.monthly_salary, emp.special)
        ta         = prorated_amount(emp.monthly_salary, emp.ta)
        da         = prorated_amount(emp.monthly_salary, emp.da)



        epfo = ((salary_obj.net_salary + salary_obj.loan_deduction_total) * emp.epfo) // 100
        final_net_salary=round(salary_obj.net_salary + salary_obj.loan_deduction_total + conveyance + hra + special + ta + da)

        esic = (final_net_salary * emp.esic) // 100

        # NEW: Professional Tax (The PDF shows a fixed 200 deduction )
        # You can change this to a dynamic field if your model has it (e.g., salary_obj.prof_tax)
        prof_tax = emp.professional_tax if emp.professional_tax else (
        0 if final_net_salary <= 3000 else
        20 if final_net_salary <= 6000 else
        80 if final_net_salary <= 9000 else
        150 if final_net_salary <= 12000 else
        200
    )


        # Handle nullable fields for Loan/Other Deductions
        # other_deductions = salary_obj.deductions if salary_obj.deductions else 0
        other_deductions = 0
        loan_deductions = salary_obj.loan_deduction_total if salary_obj.loan_deduction_total else 0
        
    
        # TOTAL DEDUCTIONS (PF + ESIC + PT + Loan + Others)
        total_deductions = epfo + esic + prof_tax + other_deductions + loan_deductions

        # 4. Final Net Pay
        # Logic Fix: Gross Earnings - Total Deductions
        final_net_pay = round(final_net_salary - total_deductions)
       












        # Final Net Salary
        salary_obj.final_net_salary = (
            salary_obj.net_salary +
            salary_obj.loan_deduction_total +
            conveyance + hra + special + ta + da
        )

      
        if  emp.monthly_salary:
            salary_obj.total_amount = (
                emp.monthly_salary +

               (emp.monthly_salary * emp.conveyance) / 100+ (emp.monthly_salary * emp.hra) / 100 + (emp.monthly_salary * emp.special) / 100 + (emp.monthly_salary * emp.ta) / 100 + (emp.monthly_salary * emp.da) / 100
            )
        else:
             salary_obj.total_amount = "-"
        salary_obj.final_net_pay=final_net_pay
        salary_obj.deductions=total_deductions

    context = {
        'salaryslip': "active",
        'data': salary_list
    }

    return render(request, "geneartedsalarylist.html", context)


   


from calendar import monthrange
from datetime import date, timedelta
from django.shortcuts import get_object_or_404
def generate_monthly_salary(employee_id, month, year):
    employee = get_object_or_404(Employee, id=employee_id)


    # Prevent duplicate salary
    if MonthlySalary.objects.filter(employee=employee, month=month, year=year).exists():
        print("Salary already generated for this month.")
        return {"error": "Salary already generated for this month."}
    

    # Salary structure
    
    try:
        structure = employee.monthly_salary
    except Salary.DoesNotExist:


        return {"error": "Salary structure not defined for employee."}

    
    


    days_in_month = monthrange(year, month)[1]

    # -------------------------------
    # COUNTERS
    # -------------------------------
    present = 0
    half = 0
    absent = 0
    paid_leave_used = 0
    unpaid_leave = 0


    from decimal import Decimal

    # Fetch only approved + not cancelled + EMI-based active loans
    active_loans = AdvanceSalaryApplication.objects.filter(
        employee=employee,
        admin_approval=True,
        superadmin_approval=True,
        admin_cancelled=False,
        superadmin_cancelled=False,
        repayment_method="emi"
    )

    loan_deduction_total = Decimal(0)

    for loan in active_loans:

    # If EMI months not set → skip safely
        if not loan.emi_months:
            continue

        # Calculate EMI amount
        emi_amount = (loan.amount_requested / loan.emi_months)

        # Add to total deduction
        loan_deduction_total += emi_amount


    if not structure:

        attendancee = Attendance.objects.filter(
        employee=employee,
        date__year=year,
        date__month=month
                    )
        present_days = attendancee.filter(type="Present").count()
        half_days = attendancee.filter(type="Half Day").count()
        total_working_days = present_days + (half_days * 0.5)
        from decimal import Decimal
        salary = Decimal(total_working_days) * employee.daily_salary
        net_salary=salary
        record = MonthlySalary.objects.create(
        employee=employee,
        month=month,
        year=year,
        total_working_days=days_in_month,
        present_days=present_days,
        half_days=half_days,
        absent_days=days_in_month - total_working_days,
        paid_leave_used=0,
        unpaid_leave=0,
        gross_salary=salary,
        deductions= - loan_deduction_total,
        net_salary=net_salary - loan_deduction_total,
        loan_deduction_total=loan_deduction_total
    )      


      
        return {"success": True, "salary": record,"message":"Salary generated successfully for Daily Wages Employee."}
   

    # -------------------------------
    # HOLIDAYS
    # -------------------------------
    holidays = Holiday.objects.filter(
        date__month=month,
        date__year=year,
        status="Active"
    ).values_list("date", flat=True)
    

    # -------------------------------
    # LOOP THROUGH DAYS
    # -------------------------------
    current_date = date(year, month, 1)

    while current_date.month == month:
        weekday = current_date.strftime("%A").lower()

        # weekly off
        is_weekly_off = False
        if employee.weekly_off:
            if employee.weekly_off.lower() in weekday:
                is_weekly_off = True

        # Skip holidays & weekly offs
        if current_date in holidays or is_weekly_off:
            current_date += timedelta(days=1)
            continue

        # Attendance fetch
        att = Attendance.objects.filter(employee=employee, date=current_date).first()

        if att:
            if att.type == "Full Day":
                present += 1
            elif att.type == "Half Day":
                half += 1
            elif att.type == "Absent":
                absent += 1
        else:
            # No attendance → Absent
            absent += 1

        current_date += timedelta(days=1)

    # -------------------------------
    # EMPTY LOGIN RULE:
    # 3 missing punch-ins = 1 half day
    # -------------------------------
    empty_login_days = Attendance.objects.filter(
        employee=employee,
        date__month=month,
        date__year=year,
        check_in__isnull=True
    ).count()

   
    extra_half_days = empty_login_days // 3
    half += extra_half_days

    # -------------------------------
    # HALF-DAY CONVERSION
    # -------------------------------
    present += (half // 2)
    print("Half days converted to full days:", half // 2)

    if half % 2 != 0:
        absent += 1  # remaining half turns into 1 full absence

    # -------------------------------
    # PAID LEAVE ADJUSTMENT
    # -------------------------------
    leaves_available = employee.leaves

    if absent > 0:
        if leaves_available >= absent:
            paid_leave_used = absent
            employee.leaves -= absent  # subtract used leaves
            absent = 0
        else:
            paid_leave_used = leaves_available
            employee.leaves = 0  # all leaves used
            unpaid_leave = absent - leaves_available
            absent = 0

    employee.save()  # update available leave balance

    # -------------------------------
    # SALARY COMPUTATION
    # -------------------------------
    
    if employee.employee_type == "monthly":
        monthly_salary = structure 
    else:
        monthly_salary = employee.daily_salary * days_in_month


    per_day_salary = monthly_salary / days_in_month

    deduction_amount = per_day_salary * unpaid_leave
    net_salary = monthly_salary - deduction_amount
    print("Net salary calculated:", net_salary)

    # -------------------------------
    # SAVE SALARY RECORD
    # -------------------------------
    record = MonthlySalary.objects.create(
        employee=employee,
        month=month,
        year=year,
        total_working_days=days_in_month,
        present_days=present,
        half_days=half,
        absent_days=absent,
        paid_leave_used=paid_leave_used,
        unpaid_leave=unpaid_leave,
        gross_salary=monthly_salary,
        deductions=deduction_amount + loan_deduction_total,
        net_salary=net_salary - loan_deduction_total,
        loan_deduction_total=loan_deduction_total,
    )


    return {"success": True, "salary": record}



def generate_salary_ajax(request):
    if request.method == "POST":
        employee_id = request.POST.get("employee_id")
        month = int(request.POST.get("month"))
        year = int(request.POST.get("year"))

        result = generate_monthly_salary(employee_id, month, year)

        if "error" in result:
            return JsonResponse({"success": False, "error": result["error"]})

        salary = result["salary"]

        return JsonResponse({
            "success": True,
            "salary_id": salary.id
        })

    return JsonResponse({"success": False, "error": "Invalid request"})


# @login_required()
# def payslipview(request,data):
#     emp=Employee.objects.get(id=MonthlySalary.objects.get(id=data).employee.id)
#     conveyance=(emp.monthly_salary * emp.conveyance) // 100
#     hra=(emp.monthly_salary * emp.hra) // 100
#     special=(emp.monthly_salary * emp.special) // 100
#     ta=(emp.monthly_salary * emp.ta) // 100
#     da=(emp.monthly_salary * emp.da) // 100
#     epfo=(emp.monthly_salary * emp.epfo) // 100
#     esic=(emp.monthly_salary * emp.esic) // 100
#     totalsalary=emp.monthly_salary+conveyance+hra+special+ta+da -epfo -esic
#     netsalary=emp.monthly_salary+conveyance+hra+special+ta+da 
#     totalsalaryn=epfo+esic+MonthlySalary.objects.get(id=data).deductions+MonthlySalary.objects.get(id=data).loan_deduction_total
#     finalnet=totalsalary - totalsalaryn
#     d={'salaryslip':"active","data": MonthlySalary.objects.get(id=data),'conveyance':conveyance,'hra':hra,'special':special,'ta':ta,'da':da,'epfo':epfo,'esic':esic,'totalsalary':totalsalary,'netsalary':netsalary,"totalsalaryn":totalsalaryn,"finalnet":finalnet}
#     return render(request,"payslip.html",d)

@login_required()
def payslipview(request, data):

    
    import calendar

    today = date.today()

    # total days in current month
    total_days_in_month = calendar.monthrange(today.year, today.month)[1]
    # 1. Fetch Data
    salary_obj = get_object_or_404(MonthlySalary, id=data)
    emp = salary_obj.employee
    
    

    # 2. Calculate Earnings (Allowances)
    # Assuming your DB stores percentages (e.g., 50 for 50%), keep // 100.
    # If DB stores flat amounts, remove // 100.
    present_days = getattr(salary_obj, 'present_days', 27)
    holidayss = len(Holiday.objects.filter(
        date__month=today.month,
        date__year=today.year,
        status="Active"
    ).values_list("date", flat=True))
    
   

    
  
    
    if emp.joining_date:
 
        joining_date = emp.joining_date

        # Check if joined in same month & year
      
        if joining_date.year == today.year and joining_date.month == today.month:
         
            join_day = joining_date.day

            if join_day <= 7:
                holidays = 4
            elif join_day <= 14:
                holidays = 3
            elif join_day <= 21:
                holidays = 2
            else:
                holidays = 1
        else:
            
              holidays = getattr(salary_obj, 'holidays', 4)
    holidays=holidays+holidayss


    total_present = present_days + holidays

    def prorated_amount(monthly_salary, percentage):
        monthly_amount = (monthly_salary * percentage) / 100
        per_day_amount = monthly_amount / total_days_in_month
        return round(per_day_amount * total_present, 2)

    # Salary components (day-wise)
    conveyance = prorated_amount(emp.monthly_salary, emp.conveyance)
    hra        = prorated_amount(emp.monthly_salary, emp.hra)
    special    = prorated_amount(emp.monthly_salary, emp.special)
    ta         = prorated_amount(emp.monthly_salary, emp.ta)
    da         = prorated_amount(emp.monthly_salary, emp.da)
    
    # GROSS EARNINGS (Basic + All Allowances)
    # In your HTML this maps to 'Total Earnings'
    gross_salary = salary_obj.net_salary + salary_obj.loan_deduction_total 
    final_net_salary = salary_obj.net_salary + salary_obj.loan_deduction_total + conveyance + hra + special + ta + da

    total_show= emp.monthly_salary + salary_obj.loan_deduction_total +round((emp.monthly_salary * emp.conveyance)/100, 2) +round((emp.monthly_salary * emp.hra)/100, 2) +round((emp.monthly_salary * emp.special)/100, 2) +round((emp.monthly_salary * emp.ta)/100, 2) +round((emp.monthly_salary * emp.da)/100, 2)

    
    # 3. Calculate Deductions
    epfo = ((salary_obj.net_salary + salary_obj.loan_deduction_total) * emp.epfo) // 100

    esic = (final_net_salary * emp.esic) // 100

    # NEW: Professional Tax (The PDF shows a fixed 200 deduction )
    # You can change this to a dynamic field if your model has it (e.g., salary_obj.prof_tax)
    prof_tax = emp.professional_tax if emp.professional_tax else (
    0 if final_net_salary <= 3000 else
    20 if final_net_salary <= 6000 else
    80 if final_net_salary <= 9000 else
    150 if final_net_salary <= 12000 else
    200
)


    # Handle nullable fields for Loan/Other Deductions
    # other_deductions = salary_obj.deductions if salary_obj.deductions else 0
    other_deductions = 0
    loan_deductions = salary_obj.loan_deduction_total if salary_obj.loan_deduction_total else 0
    
   
    # TOTAL DEDUCTIONS (PF + ESIC + PT + Loan + Others)
    total_deductions = epfo + esic + prof_tax + other_deductions + loan_deductions

    # 4. Final Net Pay
    # Logic Fix: Gross Earnings - Total Deductions
    final_net_pay = round(final_net_salary - total_deductions)
    final_net_pay_word=amount_in_words(final_net_pay)


    # 5. Context for Template
    d = {
        'salaryslip': "active",
        'data': salary_obj,
        'final_net_pay_word': final_net_pay_word,
        
        # Earnings Breakdown
        'conveyance': conveyance,
        'hra': hra,
        'special': special,
        'ta': ta,
        'da': da,
        
        # Deductions Breakdown
        'epfo': epfo,
        'esic': esic,
        'professional_tax': prof_tax, 
        'conveyance_total':round((emp.monthly_salary * emp.conveyance)/100, 2),
        'special_total':round((emp.monthly_salary * emp.special)/100, 2),
        'ta_total':round((emp.monthly_salary * emp.ta)/100, 2),
        'da_total':round((emp.monthly_salary * emp.da)/100, 2),
        'hra_total':round((emp.monthly_salary * emp.hra)/100, 2),
        'total_show':total_show,
        
        # Totals
        'netsalary': gross_salary,       # Total Earnings (Gross)
        'totalsalaryn': total_deductions, # Total Deductions
        'finalnet': final_net_pay,       # Final Pay Amount
        'employee': emp,
        'net_salary': salary_obj.net_salary + salary_obj.loan_deduction_total,
        'final_net_salary': final_net_salary,
        'total_show': total_show,
        # Attendance Details (Matches the table in PDF )
        # If your MonthlySalary model has these fields, replace the hardcoded numbers
        'attendance': {
            'present': getattr(salary_obj, 'present_days', 27), # Default 27 if field missing
            'holidays': holidays,     # Default 4 if field missing
            'absent': getattr(salary_obj, 'absent_days', 0),    # Default 0 if field missing
            'total': total_days_in_month
        }
    }
    return render(request, "payslip.html", d)


def number_to_words(n):
    ones = ["", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine",
            "Ten", "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen",
            "Sixteen", "Seventeen", "Eighteen", "Nineteen"]
    
    tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty",
            "Sixty", "Seventy", "Eighty", "Ninety"]

    def two_digits(num):
        if num < 20:
            return ones[num]
        return tens[num // 10] + (" " + ones[num % 10] if num % 10 != 0 else "")

    def three_digits(num):
        if num < 100:
            return two_digits(num)
        return ones[num // 100] + " Hundred" + (" " + two_digits(num % 100) if num % 100 != 0 else "")

    crore = n // 10000000
    n %= 10000000
    lakh = n // 100000
    n %= 100000
    thousand = n // 1000
    n %= 1000
    hundred = n

    words = ""
    if crore:
        words += three_digits(crore) + " Crore "
    if lakh:
        words += three_digits(lakh) + " Lakh "
    if thousand:
        words += three_digits(thousand) + " Thousand "
    if hundred:
        words += three_digits(hundred)

    return words.strip()


def amount_in_words(amount):
    rupees = int(amount)
    paise = int(round((amount - rupees) * 100))

    result = number_to_words(rupees) + " Rupees"
    if paise > 0:
        result += " and " + number_to_words(paise) + " Paise"
    return result + " Only"



def update_employee(request, id):
    emp = Employee.objects.get(id=id)
    

    if request.method == "POST":
 
        emp.name = request.POST.get("name")
        emp.email = request.POST.get("email")
        emp.phone = request.POST.get("phone")
        emp.gender = request.POST.get("gender")
        emp.grade = Grade.objects.get(id=request.POST.get("grade"))
        emp.position = Position.objects.get(id=request.POST.get("position"))
        emp.department = Department.objects.get(id=request.POST.get("department"))
        emp.designation = Designation.objects.get(id=request.POST.get("designation"))
        emp.birthday = request.POST.get("birthday") or None
        emp.nationality = request.POST.get("nationality")
        emp.religion = request.POST.get("religion")
        emp.marital_status = request.POST.get("marital_status")

        if "profile" in request.FILES:
            emp.profile = request.FILES["profile"]

        emp.save()
        return redirect(f"/edetails/{id}" )

    return redirect(f"/edetails/{id}" )



from datetime import datetime

def update_basic_info(request, id):
    emp = get_object_or_404(Employee, id=id)

    if request.method == "POST":
        emp.phone = request.POST.get("phone")
        emp.email = request.POST.get("email")
        emp.gender = request.POST.get("gender")

        # Parse birthday safely
        birthday_str = request.POST.get("birthday")
        if birthday_str:
            try:
                emp.birthday = datetime.strptime(birthday_str, "%Y-%m-%d").date()
            except ValueError:
                emp.birthday = None
        else:
            emp.birthday = None

        emp.save()

        return JsonResponse({
            "phone": emp.phone,
            "email": emp.email,
            "gender": emp.gender,
            "birthday": emp.birthday.strftime("%d %b %Y") if emp.birthday else "-"
        })
    
def update_personal_info(request, id):
    emp = get_object_or_404(Employee, id=id)

    if request.method == "POST":
        emp.nationality = request.POST.get("nationality")
        emp.religion = request.POST.get("religion")
        emp.marital_status = request.POST.get("marital_status")
        emp.save()

        return JsonResponse({
            "nationality": emp.nationality,
            "religion": emp.religion,
            "marital_status": emp.marital_status,
        })
    

def update_emergency(request, id):
    emp = Employee.objects.get(id=id)

    if request.method == "POST":
        emp.emergency_number = request.POST.get("emergency_number")
        emp.save()

        return JsonResponse({
            "emergency_number": emp.emergency_number
        })


def update_description(request, id):
    emp = Employee.objects.get(id=id)

    if request.method == "POST":
        emp.description = request.POST.get("description")
        emp.save()

        return JsonResponse({
            "description": emp.description
        })


def update_bank_info(request, id):
    emp = Employee.objects.get(id=id)

    if request.method == "POST":
        emp.bank_name = request.POST.get("bank_name")
        emp.account_number = request.POST.get("account_number")
        emp.ifsc_number = request.POST.get("ifsc_number")

        if "bank_passbook" in request.FILES:
            emp.bank_passbook = request.FILES["bank_passbook"]

        emp.save()

        return JsonResponse({
            "bank_name": emp.bank_name,
            "account_number": emp.account_number,
            "ifsc_number": emp.ifsc_number,
            "bank_passbook_url": emp.bank_passbook.url if emp.bank_passbook else None
        })


def update_aadhar(request, id):
    emp = Employee.objects.get(id=id)

    if request.method == "POST":
        if "aadhar_front" in request.FILES:
            emp.aadhar_front = request.FILES["aadhar_front"]
        if "aadhar_back" in request.FILES:
            emp.aadhar_back = request.FILES["aadhar_back"]
        if "pan_image" in request.FILES:
            emp.pan_image = request.FILES["pan_image"]

        emp.save()

        return JsonResponse({
            "aadhar_front_url": emp.aadhar_front.url if emp.aadhar_front else None,
            "aadhar_back_url": emp.aadhar_back.url if emp.aadhar_back else None
        })


# def update_salary(request, id):
#     emp = Employee.objects.get(id=id)

#     if request.method == "POST":
#         emp.monthly_salary = request.POST.get("monthly_salary")
#         emp.daily_salary = request.POST.get("daily_salary")
#         emp.pf_fund = request.POST.get("pf_fund")
#         emp.pf_number = request.POST.get("pf_number")
#         # update all other fields similarly
#         emp.save()

#         total_salary = calculate_total_salary(emp)  # your function to compute total

#         return JsonResponse({
#             "monthly_salary": emp.monthly_salary,
#             "daily_salary": emp.daily_salary,
#             "pf_fund": emp.pf_fund,
#             "pf_number": emp.pf_number,
#             "total_salary": total_salary
#         })

def to_decimal(value,default=0):
    return decimal.Decimal(value) if value not in [None, "", "None"] else decimal.Decimal(default)

import decimal

def update_salary(request, id):
    emp = Employee.objects.get(id=id)

    if request.method == "POST":

        emp.employee_type = request.POST.get("employee_type")

        # Salary handling
        emp.monthly_salary = to_decimal(request.POST.get("monthly_salary"))
        emp.daily_salary = to_decimal(request.POST.get("daily_salary"))

        # Other fields (all decimal/text safe)
        emp.pf_fund = to_decimal(request.POST.get("pf_fund"))
        emp.pf_number = request.POST.get("pf_number") or ""

        emp.esi_number = request.POST.get("esi_number") or ""
        emp.uan_number = request.POST.get("uan_number") or ""

        emp.leaves = to_decimal(request.POST.get("leaves"))
        emp.bonus = to_decimal(request.POST.get("bonus"))

        emp.professional_tax = to_decimal(request.POST.get("professional_tax"))
        emp.labour_tax = to_decimal(request.POST.get("labour_tax"))

        emp.weekly_off = request.POST.get("weekly_off") or ""

        emp.assign_location = request.POST.get("assign_location") or ""

        emp.conveyance = to_decimal(request.POST.get("conveyance"))
        emp.hra = to_decimal(request.POST.get("hra"))
        emp.special = to_decimal(request.POST.get("special"))
        emp.ta = to_decimal(request.POST.get("ta"))
        emp.da = to_decimal(request.POST.get("da"))

        emp.epfo = to_decimal(request.POST.get("epfo"))
        emp.esic = to_decimal(request.POST.get("esic"))
        emp.insurance = to_decimal(request.POST.get("insurance"))
        emp.pt = to_decimal(request.POST.get("pt"))

        emp.save()

        # total_salary = calculate_total_salary(emp)

        return JsonResponse({
            "status": "success",
            # "total_salary": total_salary
        })
    

from django.contrib.auth import update_session_auth_hash
from django.contrib import messages
@login_required
def change_password(request,data):
    if request.method == 'POST':

        new_password = request.POST.get('newPassword')
        confirm_password = request.POST.get('confirmPassword')

        if new_password != confirm_password:
            messages.error(request, "New password and confirm password do not match.")
            return redirect(request.META.get('HTTP_REFERER', '/'))
          

        user = User.objects.get(email=data)
    
        
      

        user.set_password(new_password)
        user.save()
        
        messages.success(request, "Password changed successfully!")
        return redirect(request.META.get('HTTP_REFERER', '/'))

    # Optional: if GET request, redirect or render a page
    return redirect('profile')




# from django.http import HttpResponse
# from django.template.loader import get_template
# from reportlab.lib.pagesizes import A4
# from reportlab.pdfgen import canvas
# from django.http import HttpResponse

# from datetime import date

# def attendance_pdf(request):
#     month = int(request.GET.get('month') or date.today().month)
#     year = int(request.GET.get('year') or date.today().year)

#     attendance_records = Attendance.objects.filter(
#         date__month=month,
#         date__year=year
#     )

#     response = HttpResponse(content_type='application/pdf')
#     response['Content-Disposition'] = 'attachment; filename="attendance.pdf"'

#     p = canvas.Canvas(response, pagesize=A4)
#     width, height = A4

#     y = height - 40
#     p.setFont("Helvetica-Bold", 14)
#     p.drawString(40, y, f"Employee Attendance - {month}/{year}")

#     y -= 30
#     p.setFont("Helvetica", 10)

#     p.drawString(40, y, "Date")
#     p.drawString(150, y, "Check In")
#     p.drawString(260, y, "Status")
#     p.drawString(350, y, "Check Out")

#     y -= 15
#     p.line(40, y, 550, y)

#     for record in attendance_records:
#         y -= 20
#         if y < 40:
#             p.showPage()
#             y = height - 40
#             p.setFont("Helvetica", 10)

#         p.drawString(40, y, record.date.strftime("%d %b %Y"))
#         p.drawString(150, y, str(record.check_in or "-"))
#         p.drawString(260, y, record.type)
#         p.drawString(350, y, str(record.check_out or "-"))

#     p.showPage()
#     p.save()
#     return response

from datetime import date
from django.http import HttpResponse
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors

# def attendance_pdf(request):
#     month = int(request.GET.get('month') or date.today().month)
#     year = int(request.GET.get('year') or date.today().year)

#     attendance_records = Attendance.objects.filter(
#         date__month=month,
#         date__year=year
#     )

#     response = HttpResponse(content_type='application/pdf')
#     response['Content-Disposition'] = 'attachment; filename="attendance.pdf"'

#     doc = SimpleDocTemplate(
#         response,
#         pagesize=A4,
#         rightMargin=30,
#         leftMargin=30,
#         topMargin=40,
#         bottomMargin=30
#     )

#     styles = getSampleStyleSheet()
#     elements = []

#     # Title (HTML-like heading)
#     elements.append(
#         Paragraph(
#             f"<b>Employee Attendance – {month}/{year}</b>",
#             styles["Title"]
#         )
#     )

#     elements.append(Paragraph("<br/>", styles["Normal"]))

#     # Table (HTML-style)
#     table_data = [
#         ["Date", "Check In", "Status", "Check Out"]
#     ]

#     for record in attendance_records:
#         table_data.append([
#             record.date.strftime("%d %b %Y"),
#             record.check_in or "-",
#             record.type,
#             record.check_out or "-"
#         ])

#     table = Table(table_data, colWidths=[120, 100, 100, 100])

#     table.setStyle(TableStyle([
#         ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
#         ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
#         ("FONT", (0, 0), (-1, 0), "Helvetica-Bold"),
#         ("ALIGN", (1, 1), (-1, -1), "CENTER"),
#         ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
#         ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
#         ("TOPPADDING", (0, 0), (-1, 0), 10),
#     ]))

#     elements.append(table)

#     doc.build(elements)
#     return response
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from datetime import date
from django.http import HttpResponse


def attendance_pdf(request):
    month = int(request.GET.get('month') or date.today().month)
    year = int(request.GET.get('year') or date.today().year)

    attendance_records = Attendance.objects.filter(
        date__month=month,
        date__year=year
    )

    employee = attendance_records.first().employee if attendance_records.exists() else None

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="attendance.pdf"'

    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        rightMargin=30,
        leftMargin=30,
        topMargin=40,
        bottomMargin=30
    )

    styles = getSampleStyleSheet()
    elements = []

    # 🔹 Title (Centered & Bold)
    title_style = ParagraphStyle(
        "TitleStyle",
        parent=styles["Title"],
        alignment=1,  # center
        fontSize=18,
        spaceAfter=20
    )

    elements.append(
        Paragraph(
            f"Employee Attendance Report<br/><font size='11'>{month}/{year}</font>",
            title_style
        )
    )

    # 🔹 Employee Info Card
    if employee:
        emp_table = Table(
            [
                ["Employee Name", employee.name],
                ["Employee ID", employee.id],
            ],
            colWidths=[120, 300]
        )

        emp_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, -1), colors.whitesmoke),
            ("FONT", (0, 0), (-1, -1), "Helvetica"),
            ("FONT", (0, 0), (0, -1), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.lightgrey),
            ("PADDING", (0, 0), (-1, -1), 8),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ]))

        elements.append(emp_table)
        elements.append(Spacer(1, 20))

    # 🔹 Attendance Table
    table_data = [["Date", "Check In", "Status", "Check Out"]]

    for record in attendance_records:
        table_data.append([
            record.date.strftime("%d %b %Y"),
            record.check_in or "-",
            record.type,
            record.check_out or "-"
        ])

    attendance_table = Table(
        table_data,
        colWidths=[120, 100, 100, 100],
        repeatRows=1
    )

    attendance_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F2F4F7")),
        ("FONT", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (1, 1), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
        ("TOPPADDING", (0, 0), (-1, 0), 10),
        ("PADDING", (0, 1), (-1, -1), 8),
    ]))

    elements.append(attendance_table)

    doc.build(elements)
    return response
